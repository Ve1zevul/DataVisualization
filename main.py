import psycopg2
import pandas as pd
import matplotlib.pyplot as plt
import os
import plotly.express as px
import numpy as np
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
DB_CONFIG = {
    'host': 'localhost',
    'database': 'DataVisualization',
    'user': 'postgres',
    'password': '0000',
    'port': '5432'
}

os.makedirs("charts", exist_ok=True)


def run_query(query):
    conn = psycopg2.connect(**DB_CONFIG)
    df = pd.read_sql(query, conn)
    conn.close()
    return df




def export_to_excel(dataframes_dict, filename):
    os.makedirs("exports", exist_ok=True)
    filepath = os.path.join("exports", filename)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        for sheet_name, df in dataframes_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(filepath)
    total_rows = 0

    for sheet_name, df in dataframes_dict.items():
        ws = wb[sheet_name]
        total_rows += len(df)

        ws.freeze_panes = "B2"

        ws.auto_filter.ref = ws.dimensions

        for col_idx, col in enumerate(df.columns, start=1):
            if pd.api.types.is_numeric_dtype(df[col]):
                col_letter = ws.cell(row=1, column=col_idx).column_letter
                data_range = f"{col_letter}2:{col_letter}{len(df)+1}"

                rule = ColorScaleRule(
                    start_type="min", start_color="FFAA0000",   # red
                    mid_type="percentile", mid_value=50, mid_color="FFFFFF00",  # yellow
                    end_type="max", end_color="FF00AA00"  # green
                )
                ws.conditional_formatting.add(data_range, rule)

    wb.save(filepath)

    print(f"Created file {filename}, {len(dataframes_dict)} sheets, {total_rows} rows")

df_orders = run_query("SELECT * FROM orders LIMIT 200;")
df_items = run_query("SELECT * FROM order_items LIMIT 200;")

export_to_excel(
    {"Orders": df_orders, "Order Items": df_items},
    "report.xlsx"
)







query = """
    SELECT
        DATE_TRUNC('month', o.order_purchase_timestamp) AS month,
        pr.product_category_name,
        SUM(oi.price + oi.freight_value) AS revenue
    FROM order_items oi
    INNER JOIN orders o ON oi.order_id = o.order_id
    INNER JOIN products pr ON oi.product_id = pr.product_id
    WHERE EXTRACT(YEAR FROM o.order_purchase_timestamp) = 2017
    GROUP BY month, pr.product_category_name
    ORDER BY month, revenue DESC;
"""

df = run_query(query)

df["month"] = pd.to_datetime(df["month"]).dt.strftime("%Y-%m")

import plotly.express as px

fig = px.bar(
    df,
    x="product_category_name",
    y="revenue",
    color="product_category_name",
    animation_frame="month",
    title="Monthly Revenue by Product Category (2017)"
)

fig.update_layout(xaxis_title="Product Category", yaxis_title="Revenue")

fig.show()




# query_pie = """
#             SELECT p.payment_type, COUNT(*) AS count
#             FROM order_payments p
#                 INNER JOIN orders o
#             ON p.order_id = o.order_id
#                 INNER JOIN customers c ON o.customer_id = c.customer_id
#             GROUP BY p.payment_type;
#             """
# df_pie = run_query(query_pie)
# df_pie.set_index("payment_type").plot.pie(y="count", autopct='%1.1f%%', legend=False)
# plt.title("Distribution of Payment Types")
# plt.ylabel("")
# plt.savefig("charts/pie_payment_types.png", bbox_inches="tight")
# plt.close()
#



# def insert_new_product():
#     conn = psycopg2.connect(**DB_CONFIG)
#     cur = conn.cursor()
#
#     # Insert a brand new product category
#     cur.execute("""
#         INSERT INTO products (product_id, product_category_name)
#         VALUES ('demo_product_new', 'holographic_gadgets');
#     """)
#
#     # Link it to an order via order_items
#     cur.execute("""
#         INSERT INTO order_items (order_id, order_item_id, product_id, seller_id, shipping_limit_date, price, freight_value)
#         VALUES ('demo_order_new', 1, 'demo_product_new', 'demo_seller', NOW(), 150.00, 20.00);
#     """)
#
#     conn.commit()
#     cur.close()
#     conn.close()
#     print("✅ Inserted new product category 'holographic_gadgets' and linked it to an order.")
# query_bar = """
#             SELECT pr.product_category_name, COUNT(*) AS total_orders
#             FROM products pr
#                      LEFT JOIN order_items oi ON pr.product_id = oi.product_id
#                      LEFT JOIN orders o ON oi.order_id = o.order_id
#             GROUP BY pr.product_category_name
#             ORDER BY total_orders ASC LIMIT 10;
#             """
# df_bar = run_query(query_bar)
# df_bar.plot.bar(x="product_category_name", y="total_orders", legend=False)
# plt.title("Top 10 Product Categories by Orders")
# plt.xlabel("Category")
# plt.ylabel("Orders")
# plt.xticks(rotation=45, ha="right")
# plt.savefig("charts/bar_top_categories.png", bbox_inches="tight")
# plt.close()
#




# query_hbar = """
#              SELECT s.seller_id, COALESCE(SUM(oi.price), 0) AS revenue
#              FROM sellers s
#                       RIGHT JOIN order_items oi ON s.seller_id = oi.seller_id
#                       RIGHT JOIN orders o ON oi.order_id = o.order_id
#              GROUP BY s.seller_id
#              ORDER BY revenue DESC LIMIT 10; \
#              """
# df_hbar = run_query(query_hbar)
# df_hbar.plot.barh(x="seller_id", y="revenue", legend=False)
# plt.title("Top 10 Sellers by Revenue")
# plt.xlabel("Revenue")
# plt.ylabel("Seller ID")
# plt.savefig("charts/hbar_top_sellers.png", bbox_inches="tight")
# plt.close()
#




# query_line = """
#              SELECT pr.product_category_name, ROUND(AVG(r.review_score), 2) AS avg_score
#              FROM products pr
#                       INNER JOIN order_items oi ON pr.product_id = oi.product_id
#                       INNER JOIN orders o ON oi.order_id = o.order_id
#                       INNER JOIN order_reviews r ON o.order_id = r.order_id
#              GROUP BY pr.product_category_name
#              ORDER BY avg_score DESC LIMIT 10;
#              """
# df_line = run_query(query_line)
#
# plt.figure(figsize=(12, 6))
# plt.plot(df_line["product_category_name"], df_line["avg_score"], marker="o")
# plt.title("Average Review Score per Product Category (Top 10)")
# plt.xlabel("Product Category")
# plt.ylabel("Average Review Score")
# plt.xticks(rotation=75, ha="right")
# plt.tight_layout()
# plt.savefig("charts/line_avg_review_score.png", bbox_inches="tight")
# plt.close()
#



# query_hist = """
#              SELECT r.review_score
#              FROM order_reviews r
#                       LEFT JOIN orders o ON r.order_id = o.order_id
#                       LEFT JOIN customers c ON o.customer_id = c.customer_id;
#              """
# df_hist = run_query(query_hist)
# df_hist["review_score"].plot.hist(bins=5, rwidth=0.8)
# plt.title("Distribution of Review Scores")
# plt.xlabel("Review Score")
# plt.ylabel("Frequency")
# plt.savefig("charts/hist_review_scores.png", bbox_inches="tight")
# plt.close()
#




# query_scatter = """
#                 SELECT oi.price, oi.freight_value
#                 FROM order_items oi
#                          INNER JOIN orders o ON oi.order_id = o.order_id
#                          INNER JOIN customers c ON o.customer_id = c.customer_id;
#                 """
# df_scatter = run_query(query_scatter)
# df_scatter.plot.scatter(x="price", y="freight_value", alpha=0.5)
# plt.title("Price vs Freight Value")
# plt.xlabel("Price")
# plt.ylabel("Freight Value")
# plt.savefig("charts/scatter_price_vs_freight.png", bbox_inches="tight")
# plt.close()
